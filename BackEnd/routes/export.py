# backend/routes/export.py - COMPLETE FIXED VERSION

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime
from bson import ObjectId
import logging
import io
import csv
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from docx import Document
from docx.shared import Inches, Pt
from mongodb.database import get_database
from routes.auth import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/export", tags=["Export"])

# ============================================================
# MODELS
# ============================================================

class ExportRequest(BaseModel):
    job_id: str
    format: str  # csv, excel, json
    include_metadata: bool = True

class BulkExportRequest(BaseModel):
    job_ids: List[str]
    format: str
    include_metadata: bool = True

class PreviewRequest(BaseModel):
    limit: int = 10

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_text(text: str) -> str:
    """Clean HTML and normalize text"""
    if not text:
        return ""
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', text)
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    # Decode HTML entities
    import html
    text = html.unescape(text)
    return text.strip()

def flatten_dataset_row(structured_data: Dict[str, Any]) -> Dict[str, Any]:
    """Flatten structured data into a flat row for CSV/Excel export"""
    flat_row = {}
    
    # Handle scalar values
    for key in ['record_id', 'title', 'description', 'content_preview', 
                'full_text', 'source_job_id', 'source_url', 'job_name',
                'extraction_timestamp', 'word_count', 'character_count']:
        flat_row[key] = structured_data.get(key, '')
    
    # Handle list fields - join with semicolons for CSV compatibility
    list_fields = ['emails', 'phone_numbers', 'addresses', 'prices', 
                   'product_names', 'categories', 'urls', 'image_urls',
                   'key_phrases', 'entities']
    
    for field in list_fields:
        value = structured_data.get(field, [])
        if isinstance(value, list):
            flat_row[field] = '; '.join(str(v) for v in value if v)
        else:
            flat_row[field] = str(value) if value else ''
    
    return flat_row

def extract_structured_dataset(content: str, metadata: Dict = None) -> Dict[str, Any]:
    """
    Extract structured data into Kaggle-quality dataset format
    """
    structured = {
        # Primary fields for dataset
        "record_id": "",
        "title": "",
        "description": "",
        "content_preview": "",
        
        # Contact information
        "emails": [],
        "phone_numbers": [],
        "addresses": [],
        
        # Business data
        "prices": [],
        "product_names": [],
        "categories": [],
        
        # Web data
        "urls": [],
        "image_urls": [],
        
        # Metadata
        "word_count": 0,
        "character_count": 0,
        "extraction_timestamp": datetime.now().isoformat(),
        
        # Analysis
        "key_phrases": [],
        "entities": [],
        
        # Full content
        "full_text": content[:10000] if content else ""
    }
    
    if not content:
        return structured
    
    # Extract title (look for common patterns)
    title_patterns = [
        r'<title>(.*?)</title>',
        r'<h1[^>]*>(.*?)</h1>',
        r'<meta property="og:title" content="([^"]+)"',
        r'class="[^"]*title[^"]*"[^>]*>(.*?)<'
    ]
    for pattern in title_patterns:
        match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
        if match:
            structured["title"] = clean_text(match.group(1))
            break
    
    # Extract description
    desc_patterns = [
        r'<meta name="description" content="([^"]+)"',
        r'<meta property="og:description" content="([^"]+)"',
        r'class="[^"]*description[^"]*"[^>]*>(.*?)<'
    ]
    for pattern in desc_patterns:
        match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
        if match:
            structured["description"] = clean_text(match.group(1))[:500]
            break
    
    # Extract emails
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    structured["emails"] = list(set(re.findall(email_pattern, content)))[:10]
    
    # Extract phone numbers
    phone_patterns = [
        r'\+?1?[-.\s]?\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}',
        r'\+?[0-9]{1,3}[-.\s]?[0-9]{3}[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}'
    ]
    phones = set()
    for pattern in phone_patterns:
        phones.update(re.findall(pattern, content))
    structured["phone_numbers"] = list(phones)[:10]
    
    # Extract URLs
    url_pattern = r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+[^\s<>"\'(){}|\\^`\[\]]*'
    structured["urls"] = list(set(re.findall(url_pattern, content)))[:20]
    
    # Extract image URLs
    img_pattern = r'<img[^>]+src=["\']([^"\']+)["\']'
    structured["image_urls"] = list(set(re.findall(img_pattern, content)))[:20]
    
    # Extract prices
    price_patterns = [
        r'\$\s?\d+(?:,\d{3})*(?:\.\d{2})?',
        r'\d+(?:,\d{3})*(?:\.\d{2})?\s?(?:USD|EUR|GBP)',
        r'price["\']?\s*[:\=]\s*["\']?(\d+(?:\.\d{2})?)'
    ]
    prices = set()
    for pattern in price_patterns:
        matches = re.findall(pattern, content, re.IGNORECASE)
        for match in matches:
            if isinstance(match, tuple):
                match = match[0]
            prices.add(match)
    structured["prices"] = list(prices)[:20]
    
    # Extract product names (ecommerce detection)
    product_selectors = [
        r'class="[^"]*product[^"]*title[^"]*"[^>]*>(.*?)<',
        r'itemprop="name"[^>]*>(.*?)<',
        r'<h1[^>]*class="[^"]*product[^"]*"[^>]*>(.*?)<'
    ]
    products = set()
    for pattern in product_selectors:
        matches = re.findall(pattern, content, re.IGNORECASE | re.DOTALL)
        for match in matches:
            clean = clean_text(match)
            if len(clean) > 3 and len(clean) < 200:
                products.add(clean)
    structured["product_names"] = list(products)[:20]
    
    # Word and character counts
    clean_content = re.sub(r'<[^>]+>', ' ', content)
    clean_content = re.sub(r'\s+', ' ', clean_content)
    structured["word_count"] = len(clean_content.split())
    structured["character_count"] = len(clean_content)
    structured["content_preview"] = clean_content[:500]
    
    # Extract key phrases (simple N-gram extraction)
    words = clean_content.lower().split()
    stopwords = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'is', 'are', 'was', 'were'}
    bigrams = []
    for i in range(len(words) - 1):
        if words[i] not in stopwords and words[i+1] not in stopwords:
            bigram = f"{words[i]} {words[i+1]}"
            if len(bigram) > 5:
                bigrams.append(bigram)
    
    from collections import Counter
    structured["key_phrases"] = [phrase for phrase, count in Counter(bigrams).most_common(10)]
    
    # Add metadata if provided
    if metadata:
        structured["source_job_id"] = metadata.get("job_id", "")
        structured["source_url"] = metadata.get("url", "")
        structured["job_name"] = metadata.get("job_name", "")
    
    return structured

# ============================================================
# EXPORT GENERATORS
# ============================================================

def generate_kaggle_csv(dataset: List[Dict[str, Any]], filename: str) -> StreamingResponse:
    """Generate Kaggle-quality CSV"""
    if not dataset:
        raise HTTPException(status_code=400, detail="No data to export")
    
    output = io.StringIO()
    
    # Define field order for consistent output
    field_order = [
        'record_id', 'title', 'description', 'content_preview', 'full_text',
        'emails', 'phone_numbers', 'urls', 'image_urls',
        'prices', 'product_names', 'key_phrases',
        'word_count', 'character_count', 'extraction_timestamp',
        'source_job_id', 'source_url', 'job_name'
    ]
    
    writer = csv.DictWriter(output, fieldnames=field_order, restval='')
    writer.writeheader()
    
    for row in dataset:
        flat_row = flatten_dataset_row(row)
        writer.writerow(flat_row)
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
    )

def generate_kaggle_excel(dataset: List[Dict[str, Any]], filename: str) -> StreamingResponse:
    """Generate Kaggle-quality Excel with multiple sheets"""
    if not dataset:
        raise HTTPException(status_code=400, detail="No data to export")
    
    wb = Workbook()
    
    # Sheet 1: Main Data
    ws_main = wb.active
    ws_main.title = "Dataset"
    
    field_order = [
        'record_id', 'title', 'description', 'content_preview',
        'emails', 'phone_numbers', 'urls', 'prices', 'product_names',
        'word_count', 'character_count', 'extraction_timestamp'
    ]
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="00ED64", end_color="00ED64", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, field in enumerate(field_order, 1):
        cell = ws_main.cell(row=1, column=col, value=field.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, row_data in enumerate(dataset, 2):
        flat_row = flatten_dataset_row(row_data)
        for col_idx, field in enumerate(field_order, 1):
            value = flat_row.get(field, '')
            if isinstance(value, str) and len(value) > 32767:
                value = value[:32767]
            cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
    
    # Auto-fit columns
    for col in range(1, len(field_order) + 1):
        max_length = 15
        for row in range(1, min(len(dataset) + 2, 100)):
            cell_value = ws_main.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws_main.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    # Sheet 2: Statistics
    ws_stats = wb.create_sheet("Statistics")
    stats_data = [
        ["Metric", "Value"],
        ["Total Records", len(dataset)],
        ["Total Emails Found", sum(len(d.get('emails', [])) for d in dataset)],
        ["Total Phone Numbers", sum(len(d.get('phone_numbers', [])) for d in dataset)],
        ["Total URLs", sum(len(d.get('urls', [])) for d in dataset)],
        ["Total Prices", sum(len(d.get('prices', [])) for d in dataset)],
        ["Total Products", sum(len(d.get('product_names', [])) for d in dataset)],
        ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    
    for row_idx, row in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )

def generate_kaggle_json(dataset: List[Dict[str, Any]], filename: str) -> StreamingResponse:
    """Generate Kaggle-quality JSON with metadata"""
    if not dataset:
        raise HTTPException(status_code=400, detail="No data to export")
    
    output_data = {
        "metadata": {
            "export_timestamp": datetime.now().isoformat(),
            "total_records": len(dataset),
            "format_version": "1.0",
            "description": "Kaggle-quality dataset from web scraping"
        },
        "fields": list(flatten_dataset_row(dataset[0]).keys()),
        "data": [flatten_dataset_row(row) for row in dataset]
    }
    
    output = io.BytesIO()
    json_str = json.dumps(output_data, indent=2, default=str)
    output.write(json_str.encode('utf-8'))
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}.json"}
    )

# ============================================================
# API ENDPOINTS
# ============================================================

@router.get("/jobs-with-results")
async def get_jobs_with_results(
    current_user: dict = Depends(get_current_user)
):
    """Get all jobs that have parsed results available"""
    db = await get_database()
    user_id = str(current_user.get("id") or current_user.get("_id"))
    
    pipeline = [
        {"$match": {"user_id": user_id}},
        {"$lookup": {
            "from": "parsed_results",
            "localField": "_id",
            "foreignField": "job_id",
            "as": "parsed_results"
        }},
        {"$match": {"parsed_results": {"$ne": []}}},
        {"$project": {
            "name": 1,
            "url": 1,
            "status": 1,
            "created_at": 1,
            "parsed_count": {"$size": "$parsed_results"}
        }},
        {"$sort": {"created_at": -1}}
    ]
    
    cursor = db.jobs.aggregate(pipeline)
    jobs = []
    async for job in cursor:
        jobs.append({
            "id": str(job["_id"]),
            "name": job.get("name", ""),
            "url": job.get("url", ""),
            "status": job.get("status", ""),
            "created_at": job.get("created_at").isoformat() if job.get("created_at") else None,
            "parsed_count": job.get("parsed_count", 0)
        })
    
    return {"jobs": jobs}

@router.post("/preview/{job_id}")
async def preview_export(
    job_id: str,
    request: PreviewRequest,
    current_user: dict = Depends(get_current_user)
):
    """Preview parsed results before export"""
    db = await get_database()
    user_id = str(current_user.get("id") or current_user.get("_id"))
    
    if not ObjectId.is_valid(job_id):
        raise HTTPException(status_code=400, detail="Invalid job ID format")
    
    # Get job
    job = await db.jobs.find_one({
        "_id": ObjectId(job_id),
        "user_id": user_id
    })
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get parsed results with limit
    cursor = db.parsed_results.find({"job_id": ObjectId(job_id)}).limit(request.limit)
    parsed_results = []
    async for result in cursor:
        parsed_results.append(result)
    
    if not parsed_results:
        raise HTTPException(status_code=404, detail="No parsed results found")
    
    # Build preview dataset
    preview_data = []
    fields = set()
    
    for result in parsed_results:
        content = result.get("parsed_content", "")
        metadata = {
            "job_id": job_id,
            "url": job.get("url", ""),
            "job_name": job.get("name", "")
        }
        structured = extract_structured_dataset(content, metadata)
        structured["record_id"] = str(result.get("_id", ""))
        
        # Flatten for preview
        flat_row = flatten_dataset_row(structured)
        preview_data.append(flat_row)
        fields.update(flat_row.keys())
    
    # Get total count
    total_available = await db.parsed_results.count_documents({"job_id": ObjectId(job_id)})
    
    return {
        "job_id": job_id,
        "job_name": job.get("name"),
        "preview": preview_data,
        "fields": sorted(list(fields)),
        "total_available": total_available,
        "preview_limit": request.limit
    }

@router.post("/generate")
async def generate_export(
    request: ExportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Generate Kaggle-quality dataset export"""
    db = await get_database()
    user_id = str(current_user.get("id") or current_user.get("_id"))
    
    if not ObjectId.is_valid(request.job_id):
        raise HTTPException(status_code=400, detail="Invalid job ID format")
    
    # Get job
    job = await db.jobs.find_one({
        "_id": ObjectId(request.job_id),
        "user_id": user_id
    })
    
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get parsed results
    cursor = db.parsed_results.find({"job_id": ObjectId(request.job_id)})
    parsed_results = []
    async for result in cursor:
        parsed_results.append(result)
    
    if not parsed_results:
        raise HTTPException(status_code=404, detail="No parsed results found")
    
    # Build Kaggle-quality dataset
    dataset = []
    for result in parsed_results:
        content = result.get("parsed_content", "")
        metadata = {
            "job_id": request.job_id,
            "url": job.get("url", ""),
            "job_name": job.get("name", "")
        }
        structured = extract_structured_dataset(content, metadata)
        structured["record_id"] = str(result.get("_id", ""))
        dataset.append(structured)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_name = job.get("name", "export")[:30].replace(" ", "_")
    filename = f"kaggle_dataset_{job_name}_{timestamp}"
    
    format_lower = request.format.lower()
    
    if format_lower == "csv":
        return generate_kaggle_csv(dataset, filename)
    elif format_lower in ["excel", "xlsx"]:
        return generate_kaggle_excel(dataset, filename)
    elif format_lower == "json":
        return generate_kaggle_json(dataset, filename)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {request.format}")

@router.post("/bulk")
async def bulk_export(
    request: BulkExportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Export multiple jobs as a single dataset"""
    db = await get_database()
    user_id = str(current_user.get("id") or current_user.get("_id"))
    
    all_dataset = []
    
    for job_id in request.job_ids:
        if not ObjectId.is_valid(job_id):
            continue
        
        job = await db.jobs.find_one({"_id": ObjectId(job_id), "user_id": user_id})
        if not job:
            continue
        
        cursor = db.parsed_results.find({"job_id": ObjectId(job_id)})
        async for result in cursor:
            content = result.get("parsed_content", "")
            metadata = {
                "job_id": job_id,
                "url": job.get("url", ""),
                "job_name": job.get("name", "")
            }
            structured = extract_structured_dataset(content, metadata)
            structured["record_id"] = str(result.get("_id", ""))
            all_dataset.append(structured)
    
    if not all_dataset:
        raise HTTPException(status_code=404, detail="No data found for selected jobs")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"kaggle_dataset_bulk_{len(request.job_ids)}jobs_{timestamp}"
    
    format_lower = request.format.lower()
    
    if format_lower == "csv":
        return generate_kaggle_csv(all_dataset, filename)
    elif format_lower in ["excel", "xlsx"]:
        return generate_kaggle_excel(all_dataset, filename)
    elif format_lower == "json":
        return generate_kaggle_json(all_dataset, filename)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {request.format}")

@router.get("/stats/{job_id}")
async def get_export_stats(
    job_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get dataset statistics"""
    db = await get_database()
    user_id = str(current_user.get("id") or current_user.get("_id"))
    
    if not ObjectId.is_valid(job_id):
        raise HTTPException(status_code=400, detail="Invalid job ID format")
    
    job = await db.jobs.find_one({"_id": ObjectId(job_id), "user_id": user_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    cursor = db.parsed_results.find({"job_id": ObjectId(job_id)})
    results = []
    async for result in cursor:
        results.append(result)
    
    if not results:
        raise HTTPException(status_code=404, detail="No parsed results found")
    
    # Calculate statistics across all parsed results
    total_emails = 0
    total_phones = 0
    total_urls = 0
    total_prices = 0
    
    for result in results:
        content = result.get("parsed_content", "")
        structured = extract_structured_dataset(content)
        total_emails += len(structured.get("emails", []))
        total_phones += len(structured.get("phone_numbers", []))
        total_urls += len(structured.get("urls", []))
        total_prices += len(structured.get("prices", []))
    
    return {
        "job_name": job.get("name"),
        "job_url": job.get("url"),
        "total_parsed_records": len(results),
        "total_emails_found": total_emails,
        "total_phones_found": total_phones,
        "total_urls_found": total_urls,
        "total_prices_found": total_prices,
        "available_formats": ["csv", "excel", "json"],
        "estimated_records": len(results) * 10,
        "last_parsed_date": results[0].get("created_at").isoformat() if results[0].get("created_at") else None
    }